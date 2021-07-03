from openpyxl import Workbook
from openpyxl import load_workbook

class Excel():
    def __init__(self,fileName):
        self.fileName = fileName
        self.out = Workbook()
        self.fout = self.out.active
        self.fout.title = 'Sheet1'
        self.theFile = load_workbook(fileName)
        self.fin = self.theFile.active
        self.maxRow = self.fin.max_row + 1
        self.max_column = self.fin.max_column

    def add(self,paramValue):
        for i in range(1,self.maxRow):
            for j in range(0,self.max_column):
                cell_name = '{}{}'.format(chr(ord("A")+j),i)
                self.fout[cell_name] = self.fin[cell_name].value
        for j in range(0,self.max_column):
            cell_name = '{}{}'.format(chr(ord("A") + j), self.maxRow)
            self.fout[cell_name] = paramValue[j]

        self.theFile.close()
        self.out.save(self.fileName)
        self.out.close()

# # fileName = '1.xlsx'
# # out = Workbook()
# # fout = out.active
# # fout.title = 'Sheet1'
# # theFile = load_workbook(fileName)
# # fin = theFile.active
#
# maxRow = fin.max_row + 1
# max_colum = fin.max_column
# for i in range(1, maxRow):
#     for j in range(0, max_colum):
#         cell_name = '{}{}'.format(chr(ord("A") + j), i)
#         fout[cell_name] = fin[cell_name].value
#
paramValue = ['xxx', '2011年01月20日 10：00-13：00', 'xx', 'xxx', 'xx', '自动分析方式', '399', 'dssdfs.txt']
# i = maxRow
# for j in range(0,max_colum):
#     cell_name = '{}{}'.format(chr(ord("A") + j), i)
#     fout[cell_name] = paramValue[j]
#
# theFile.close()
# out.save('1.xlsx')
# out.close()
#

excel = Excel('1.xlsx')
excel.add(paramValue)
