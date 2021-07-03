from openpyxl import Workbook
from openpyxl import load_workbook

class WriteReaderExcel():
    def __init__(self, filePath):
        self.filePath = filePath
        self.wb = load_workbook(self.filePath)
        self.workSheet = self.wb.active
        self.maxRow = self.workSheet.max_row
        print(self.workSheet.max_column)
    def WriteStep2(self, header = [], footer = [],timeRange = [],paramter=[], classValue=[]):
        self.maxRow += 1
        self.workSheet.insert_rows(self.maxRow)
        idx = 0
        for i in range(0,2):
            self.workSheet[self.maxRow][idx].value = str(header[i])
            idx += 1
        for i in range(0,2):
            self.workSheet[self.maxRow][idx].value = str(footer[i])
            idx += 1
        self.workSheet[self.maxRow][idx].value = str(paramter[0])
        idx += 1
        for i in range(0, 2):
            self.workSheet[self.maxRow][idx].value = str(timeRange[i])
            idx += 1
        for i in range(2,8):
            self.workSheet[self.maxRow][idx].value = str(paramter[i])
            idx += 1
        for i in range(0,9):
            try:
                self.workSheet[self.maxRow][idx].value = str(classValue[i])
            except Exception as e:
                print(idx,i)
                print(e)
            idx += 1
        self.wb.save(self.filePath)
        self.wb.close()

# filePath = '../File/record/2.xlsx'
# excel = WriteReaderExcel(filePath)
# header = ['复旦大学','asdasda']
# footer = ['20','50']
# timeRange = ['2020.01.20 11:26','2021.07.01 20:37']
# paramter = [
#     '王炎文',
#     '',
#     '烧结矿',
#     'dsdsad-2',
#     '课题',
#     '自动分析方式',
#     '255',
#     'asdasdasd'
# ]
# classValue = [0.12,0.23,0.34,0.21,0.45,0.56,0.7,0.8,0.9]
# excel.Write(header=header,footer=footer,timeRange=timeRange,paramter=paramter,classValue=classValue)
#
# # wb = load_workbook('1.xlsx')
# # self.workSheet = wb.active
# # wb.save('1.xlsx')
# # wb.close()
